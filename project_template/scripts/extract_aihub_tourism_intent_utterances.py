from __future__ import annotations

from collections import defaultdict
import json
from pathlib import Path
import re
import sys
import tempfile
import zipfile
from typing import Any, Iterable

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

AIHUB_ROOT = PROJECT_ROOT / "data" / "external" / "aihub" / "raw"
TRAIN_OUTPUT = PROJECT_ROOT / "data" / "eval" / "tourism_intent_aihub_utterances.jsonl"
HOLDOUT_OUTPUT = PROJECT_ROOT / "data" / "eval" / "tourism_intent_aihub_holdout.jsonl"

INTENT_LIMITS = {
    "add_condition": 900,
    "recommend_places": 900,
    "unsupported_request": 900,
}
HOLDOUT_LIMITS = {
    "add_condition": 250,
    "recommend_places": 250,
    "unsupported_request": 250,
}


def clean_text(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"^[A-Z]\.\s*", "", text)
    text = re.sub(r"#@\w+#", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def is_useful_user_text(text: str) -> bool:
    if len(text) < 6 or len(text) > 120:
        return False
    if text in {"네", "아니요", "감사합니다", "알겠습니다", "예", "아니", "맞습니다"}:
        return False
    return any(char.isalpha() for char in text)


def classify_aihub_text(text: str, *, domain: str = "") -> tuple[str | None, str]:
    normalized = text.replace(" ", "")
    if any(keyword in normalized for keyword in ["감사", "덕분", "찾았네요", "추천받", "추천을받", "추천많이받"]):
        return None, ""
    unsupported_keywords = [
        "가격",
        "요금",
        "얼마",
        "예약",
        "취소",
        "환불",
        "결제",
        "입금",
        "영업",
        "운영시간",
        "문닫",
        "열었",
        "마감",
        "체크인",
        "체크아웃",
        "객실",
        "룸",
        "침구",
        "비품",
        "프론트",
        "세탁",
        "문의",
        "방이",
        "빈방",
        "좌석",
        "남았",
        "재고",
        "배송",
        "수강료",
        "진료",
        "병원",
        "약국",
        "상담",
        "몇번버스",
        "가는버스",
        "버스로갈",
        "대중교통으로가는",
        "가는방법",
        "갈아타",
        "우울",
        "불안",
        "화가",
        "죽겠",
    ]
    condition_keywords = [
        "주차",
        "화장실",
        "엘리베이터",
        "휠체어",
        "유모차",
        "유아차",
        "아이",
        "어린이",
        "노약자",
        "어르신",
        "반려견",
        "반려동물",
        "장애인",
        "경사로",
        "계단",
        "대중교통",
        "버스",
        "지하철",
        "실내",
        "비올",
        "비 오는",
    ]
    place_terms = ["곳", "펜션", "호텔", "민박", "모텔", "관광지", "명소", "숙소", "캠핑장", "업체", "장소", "볼거리"]

    if any(keyword in normalized for keyword in unsupported_keywords):
        return "unsupported_request", "unsupported_keyword"
    if any(keyword in normalized for keyword in condition_keywords) and any(
        keyword in normalized for keyword in ["가능", "있", "되", "편", "문의", "알려", "추천", "곳"]
    ):
        return "add_condition", "condition_keyword"
    recommendation_request = (
        "추천" in normalized
        or "갈만" in normalized
        or "가볼만" in normalized
        or "볼거리" in normalized
        or "명소" in normalized
        or "코스" in normalized
        or (any(keyword in normalized for keyword in ["근처", "주변", "가까운"]) and any(term in normalized for term in place_terms))
        or (any(keyword in normalized for keyword in ["찾고", "찾는데", "찾으", "찾을"]) and any(term in normalized for term in place_terms))
    )
    if domain == "tourism" and recommendation_request:
        return "recommend_places", "tourism_recommend_keyword"
    if any(keyword in normalized for keyword in ["맛집", "숙소", "펜션", "호텔"]) and any(
        keyword in normalized for keyword in ["추천", "근처", "주변", "어디", "찾"]
    ):
        return "recommend_places", "place_recommend_keyword"
    return None, ""


def iter_purpose_dialogue(zip_path: Path, *, split: str) -> Iterable[dict[str, str]]:
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.endswith(".json"):
                continue
            payload = json.loads(zf.read(name).decode("utf-8"))
            for info in payload.get("info") or []:
                annotations = info.get("annotations") or {}
                subject = str(annotations.get("subject") or "")
                for line in annotations.get("lines") or []:
                    speaker = (line.get("speaker") or {}).get("id")
                    if speaker != "B":
                        continue
                    text = clean_text(line.get("norm_text") or line.get("text"))
                    if not is_useful_user_text(text):
                        continue
                    intent, reason = classify_aihub_text(text, domain="tourism")
                    if intent:
                        yield {
                            "text": text,
                            "intent": intent,
                            "source": "aihub_purpose_dialogue",
                            "split": split,
                            "item_id": f"{Path(name).name}:{line.get('id')}",
                            "aihub_subject": subject,
                            "mapping_reason": reason,
                        }


def iter_korean_dialogue(zip_path: Path) -> Iterable[dict[str, str]]:
    with zipfile.ZipFile(zip_path) as zf:
        for name in zf.namelist():
            if not name.endswith(".xlsx"):
                continue
            with zf.open(name) as src, tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
                temp.write(src.read())
                temp_path = Path(temp.name)
            try:
                workbook = load_workbook(temp_path, read_only=True, data_only=True)
                sheet = workbook[workbook.sheetnames[0]]
                rows = sheet.iter_rows(values_only=True)
                header = [str(value or "") for value in next(rows)]
                index = {column: i for i, column in enumerate(header)}
                for row_index, row in enumerate(rows, start=2):
                    speaker = str(row[index.get("SPEAKER", -1)] or "")
                    qa = str(row[index.get("QA", -1)] or "")
                    if speaker != "고객" or qa != "Q":
                        continue
                    text = clean_text(row[index.get("MQ", -1)] or row[index.get("SENTENCE", -1)])
                    if not is_useful_user_text(text):
                        continue
                    intent, reason = classify_aihub_text(text)
                    if intent != "unsupported_request":
                        continue
                    yield {
                        "text": text,
                        "intent": intent,
                        "source": "aihub_korean_dialogue",
                        "split": "train",
                        "item_id": f"{name}:{row_index}",
                        "aihub_subject": str(row[index.get("DOMAIN", -1)] or ""),
                        "mapping_reason": reason,
                    }
            finally:
                temp_path.unlink(missing_ok=True)


def iter_wellness_dialogue(path: Path) -> Iterable[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    header = [str(value or "") for value in next(rows)]
    if "유저" in header:
        text_index = header.index("유저")
        category_index = header.index("구분")
    else:
        text_index = header.index("utterance")
        category_index = header.index("intent")
    for row_index, row in enumerate(rows, start=2):
        text = clean_text(row[text_index])
        if not is_useful_user_text(text):
            continue
        intent, reason = classify_aihub_text(text)
        if intent != "unsupported_request":
            continue
        yield {
            "text": text,
            "intent": intent,
            "source": "aihub_wellness_dialogue",
            "split": "train",
            "item_id": f"{path.name}:{row_index}",
            "aihub_subject": str(row[category_index] or ""),
            "mapping_reason": reason,
        }


def dedupe_and_limit(rows: Iterable[dict[str, str]], limits: dict[str, int]) -> list[dict[str, str]]:
    seen = set()
    counts: dict[str, int] = defaultdict(int)
    result = []
    for row in rows:
        key = (row["text"], row["intent"])
        intent = row["intent"]
        if key in seen or counts[intent] >= limits.get(intent, 0):
            continue
        seen.add(key)
        counts[intent] += 1
        result.append(row)
    return result


def write_jsonl(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False) + "\n")


def main() -> None:
    purpose_root = AIHUB_ROOT / "purpose_dialogue" / "01.데이터"
    train_rows = []
    train_rows.extend(
        iter_purpose_dialogue(
            purpose_root / "1.Training" / "라벨링데이터" / "TL_4.tourism.zip",
            split="train",
        )
    )
    train_rows.extend(iter_korean_dialogue(AIHUB_ROOT / "korean_dialogue" / "01_dialog" / "한국어대화_new_260226.zip"))
    for path in sorted((AIHUB_ROOT / "wellness_dialogue").glob("*.xlsx")):
        train_rows.extend(iter_wellness_dialogue(path))

    holdout_rows = list(
        iter_purpose_dialogue(
            purpose_root / "2.Validation" / "라벨링데이터" / "VL_4.tourism.zip",
            split="holdout",
        )
    )

    train_result = dedupe_and_limit(train_rows, INTENT_LIMITS)
    holdout_result = dedupe_and_limit(holdout_rows, HOLDOUT_LIMITS)
    write_jsonl(TRAIN_OUTPUT, train_result)
    write_jsonl(HOLDOUT_OUTPUT, holdout_result)

    def counts(rows: list[dict[str, str]]) -> dict[str, int]:
        output: dict[str, int] = defaultdict(int)
        for row in rows:
            output[row["intent"]] += 1
        return dict(sorted(output.items()))

    print(f"Wrote {len(train_result)} rows to {TRAIN_OUTPUT.relative_to(PROJECT_ROOT)}")
    print(json.dumps(counts(train_result), ensure_ascii=False, sort_keys=True))
    print(f"Wrote {len(holdout_result)} rows to {HOLDOUT_OUTPUT.relative_to(PROJECT_ROOT)}")
    print(json.dumps(counts(holdout_result), ensure_ascii=False, sort_keys=True))


if __name__ == "__main__":
    main()
