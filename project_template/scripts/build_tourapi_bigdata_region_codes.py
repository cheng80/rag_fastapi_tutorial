from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = (
    PROJECT_ROOT
    / "data"
    / "source"
    / "tourapi_guides"
    / "hub"
    / "한국관광공사_TourAPI_관광지_시군구_코드정보_v1.0.xlsx"
)
DEFAULT_OUTPUT = PROJECT_ROOT / "data" / "processed" / "tourapi_bigdata_region_codes.json"


def normalize_region_name(value: Any) -> str:
    return str(value or "").strip()


def strip_administrative_suffix(area_name: str) -> str:
    replacements = {
        "서울특별시": "서울",
        "부산광역시": "부산",
        "대구광역시": "대구",
        "인천광역시": "인천",
        "광주광역시": "광주",
        "대전광역시": "대전",
        "울산광역시": "울산",
        "세종특별자치시": "세종",
        "경기도": "경기",
        "강원특별자치도": "강원",
        "충청북도": "충북",
        "충청남도": "충남",
        "전북특별자치도": "전북",
        "전라남도": "전남",
        "경상북도": "경북",
        "경상남도": "경남",
        "제주특별자치도": "제주",
    }
    return replacements.get(area_name, area_name)


def build_region_codes(input_path: Path) -> dict[str, Any]:
    workbook = load_workbook(input_path, data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"empty workbook: {input_path}")
    headers = [normalize_region_name(value) for value in rows[0]]
    index = {header: position for position, header in enumerate(headers)}
    required = ["areaCd", "areaNm", "sigunguCd", "sigunguNm"]
    missing = [header for header in required if header not in index]
    if missing:
        raise ValueError(f"missing columns in {input_path}: {', '.join(missing)}")

    region_index: dict[str, dict[str, str]] = {}
    region_group_index: dict[str, dict[str, Any]] = {}
    area_index: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        area_cd = normalize_region_name(row[index["areaCd"]])
        area_name = normalize_region_name(row[index["areaNm"]])
        signgu_cd = normalize_region_name(row[index["sigunguCd"]])
        signgu_name = normalize_region_name(row[index["sigunguNm"]])
        if not area_cd or not area_name or not signgu_cd or not signgu_name:
            continue
        short_area_name = strip_administrative_suffix(area_name)
        area_entry = area_index.setdefault(
            area_name,
            {
                "area_cd": area_cd,
                "area_name": area_name,
                "short_area_name": short_area_name,
                "sigungu": {},
            },
        )
        area_entry["sigungu"][signgu_name] = signgu_cd
        payload = {
            "area_cd": area_cd,
            "area_name": area_name,
            "short_area_name": short_area_name,
            "signgu_cd": signgu_cd,
            "signgu_name": signgu_name,
        }
        aliases = {
            signgu_name,
            f"{area_name} {signgu_name}",
            f"{short_area_name} {signgu_name}",
        }
        if signgu_name.endswith(("시", "군", "구")):
            aliases.add(signgu_name[:-1])
        for alias in aliases:
            if alias:
                region_index.setdefault(alias, payload)
        if " " in signgu_name:
            parent_name = signgu_name.split(" ", 1)[0]
            if parent_name.endswith("시"):
                for group_alias in {parent_name, parent_name[:-1], f"{area_name} {parent_name}", f"{short_area_name} {parent_name}"}:
                    if not group_alias:
                        continue
                    group = region_group_index.setdefault(
                        group_alias,
                        {
                            "area_cd": area_cd,
                            "area_name": area_name,
                            "short_area_name": short_area_name,
                            "parent_signgu_name": parent_name,
                            "signgu_codes": [],
                        },
                    )
                    if signgu_cd not in group["signgu_codes"]:
                        group["signgu_codes"].append(signgu_cd)

    try:
        source = str(input_path.relative_to(PROJECT_ROOT))
    except ValueError:
        source = str(input_path)
    return {
        "source": source,
        "code_system": "TourAPI bigdata areaCd/signguCd",
        "area_index": area_index,
        "region_index": region_index,
        "region_group_index": region_group_index,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Build TourAPI bigdata areaCd/signguCd region code cache.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    payload = build_region_codes(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(
        f"wrote {args.output} "
        f"areas={len(payload['area_index'])} aliases={len(payload['region_index'])} "
        f"group_aliases={len(payload['region_group_index'])}"
    )


if __name__ == "__main__":
    main()
